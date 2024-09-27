import argparse
import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

TERMINOLOGY_NAME = 'excel-translator-terminology'

def translate_text(text, translate_client, source_language_code, target_language_code, terminology_names):
    try:
        response = translate_client.translate_text(
            Text=text,
            SourceLanguageCode=source_language_code,
            TargetLanguageCode=target_language_code,
            TerminologyNames=terminology_names
        )
        return response.get('TranslatedText')
    except ClientError as client_error:
        print(f'Translation error: {client_error}')
        return text

def copy_cell_format(source_cell, target_cell):
    # Copy font
    target_cell.font = Font(
        name=source_cell.font.name,
        size=source_cell.font.size,
        bold=source_cell.font.bold,
        italic=source_cell.font.italic,
        color=source_cell.font.color
    )
    
    # Copy alignment
    target_cell.alignment = Alignment(
        horizontal=source_cell.alignment.horizontal,
        vertical=source_cell.alignment.vertical,
        wrap_text=source_cell.alignment.wrap_text
    )
    
    # Copy fill
    if source_cell.fill.fill_type:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            fgColor=source_cell.fill.fgColor,
            bgColor=source_cell.fill.bgColor
        )
    
    # Copy border
    if source_cell.border:
        target_cell.border = Border(
            left=Side(border_style=source_cell.border.left.style, color=source_cell.border.left.color),
            right=Side(border_style=source_cell.border.right.style, color=source_cell.border.right.color),
            top=Side(border_style=source_cell.border.top.style, color=source_cell.border.top.color),
            bottom=Side(border_style=source_cell.border.bottom.style, color=source_cell.border.bottom.color)
        )

def translate_workbook(input_path, output_path, source_language_code, target_language_code, terminology_names, translate_client):
    # Load the original workbook
    original_wb = load_workbook(input_path, data_only=True)
    
    # Create a new workbook
    new_wb = Workbook()
    new_wb.remove(new_wb.active)  # Remove the default sheet
    
    for sheet_name in original_wb.sheetnames:
        print(f'Translating sheet: {sheet_name}')
        original_sheet = original_wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)
        
        # Copy column dimensions
        for column in original_sheet.column_dimensions:
            new_sheet.column_dimensions[column] = original_sheet.column_dimensions[column]
        
        for row_index, row in enumerate(original_sheet.iter_rows(), 1):
            for col_index, cell in enumerate(row, 1):
                new_cell = new_sheet.cell(row=row_index, column=col_index)
                
                # Copy cell format
                copy_cell_format(cell, new_cell)
                
                # Translate and set value
                if isinstance(cell.value, str) and cell.value.strip():
                    new_cell.value = translate_text(
                        cell.value,
                        translate_client,
                        source_language_code,
                        target_language_code,
                        terminology_names
                    )
                else:
                    new_cell.value = cell.value
    
    new_wb.save(output_path)

def import_terminology(terminology_file_path, translate_client):
    print(f'Importing terminology data from {terminology_file_path}...')
    with open(terminology_file_path, 'rb') as f:
        translate_client.import_terminology(
            Name=TERMINOLOGY_NAME,
            MergeStrategy='OVERWRITE',
            TerminologyData={'File': bytearray(f.read()), 'Format': 'CSV'}
        )

def main():
    argument_parser = argparse.ArgumentParser(
        'Translates Excel files from source language to target language using Amazon Translate service'
    )
    argument_parser.add_argument('source_language_code', type=str, help='Source language code (e.g., en)')
    argument_parser.add_argument('target_language_code', type=str, help='Target language code (e.g., es)')
    argument_parser.add_argument('input_file_path', type=str, help='Path to the input Excel file')
    argument_parser.add_argument('--terminology', type=str, help='Path to the terminology CSV file')
    argument_parser.add_argument('--region', type=str, required=True, help='AWS region (e.g., us-west-2)')
    args = argument_parser.parse_args()

    translate_client = boto3.client(service_name='translate', region_name=args.region)

    terminology_names = []
    if args.terminology:
        import_terminology(args.terminology, translate_client)
        terminology_names = [TERMINOLOGY_NAME]

    print(f'Translating {args.input_file_path} from {args.source_language_code} to {args.target_language_code}...')
    
    output_file_path = args.input_file_path.replace('.xlsx', f'-{args.target_language_code}.xlsx')
    translate_workbook(args.input_file_path, output_file_path, args.source_language_code, args.target_language_code, terminology_names, translate_client)
    
    print(f'Translation completed. Output saved to {output_file_path}')

if __name__ == '__main__':
    main()